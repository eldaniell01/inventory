from PyQt6 import uic
import os
from openpyxl import load_workbook
from PyQt6.QtWidgets import QApplication, QMainWindow, QWidget, QMessageBox, QFileDialog, QTableWidget, QHeaderView, QTableWidgetItem
from PyQt6.QtGui import QFont
from db.conexion import ConexionMysql
from db.querys import Query

class Home(QMainWindow):
    def __init__(self):
        super().__init__()
        self.main = uic.loadUi('views/home.ui')
        self.main.show()
        self.main.btnProveedor.clicked.connect(self.registerProveedor)
        self.main.btnMarca.clicked.connect(self.registerMarca)        
        self.main.btnClean.clicked.connect(self.cleanTables)
        self.main.btnClean2.clicked.connect(self.cleanTables)
        self.main.btnList.clicked.connect(self.openExcel)
        self.main.btnShop.clicked.connect(self.registerProduct)
        self.showTProveedor()
        self.showTMarca()
        self.showTcompra()
        self.proveedor = []
        self.showProveedor()
        
    def showTProveedor(self):
        columns = ['NIT', 'PROVEEDOR']
        self.main.tProveedor.setFont(QFont("FiraCode Nerd Font", 12))
        self.main.tProveedor.setColumnCount(len(columns))
        for column, name in enumerate(columns):
            self.main.tProveedor.setHorizontalHeaderItem(column, QTableWidgetItem(name))
        self.main.tProveedor.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        header_style = """
        QHeaderView::section {
            font-family: "FiraCode Nerd Font";
            font-size: 12pt;
            font-weight: bold;
            background-color: rgb(255, 255, 255);
        }
        QTableWidget{
            background-color: rgb(255, 255, 255);
        }
        """
        self.main.tProveedor.setStyleSheet(header_style)
    
    def showTMarca(self):
        columns = ['MARCA', 'PROVEEDOR']
        self.main.tMarca.setFont(QFont("FiraCode Nerd Font", 12))
        self.main.tMarca.setColumnCount(len(columns))
        for column, name in enumerate(columns):
            self.main.tMarca.setHorizontalHeaderItem(column, QTableWidgetItem(name))
        self.main.tMarca.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        header_style = """
        QHeaderView::section {
            font-family: "FiraCode Nerd Font";
            font-size: 12pt;
            font-weight: bold;
            background-color: rgb(255, 255, 255);
        }
        QTableWidget{
            background-color: rgb(255, 255, 255);
        }
        """
        self.main.tMarca.setStyleSheet(header_style)
        
    def showTcompra(self):
        columns = ['CANTIDAD', 'CODIGO', 'DESCRIPCION', 'MARCA','MEDIDA', 'PRECIO C.', 'PRECIO V.', 'DESCUENTO', 'SUBTOTAL' ]
        self.main.tShop.setFont(QFont("FiraCode Nerd Font", 11))
        self.main.tShop.setColumnCount(len(columns))
        for column, name in enumerate(columns):
            self.main.tShop.setHorizontalHeaderItem(column, QTableWidgetItem(name))
        self.main.tShop.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        header_style = """
        QHeaderView::section {
            font-family: "FiraCode Nerd Font";
            font-size: 12pt;
            font-weight: bold;
            background-color: rgb(255, 255, 255);
        }
        QTableWidget{
            background-color: rgb(255, 255, 255);
        }
        """
        self.main.tShop.setStyleSheet(header_style)    
    
    def showProveedor(self):
        query = Query()
        options = set(self.proveedor)
        result = query.selectProveedor()
        for id, data in result:
            self.main.cbProveedor.addItem(str(data))
            self.main.cbProveedor_2.addItem(str(data))
    
    def registerDetail(self):
        query = Query()
        fecha = self.main.dFecha.date().toString("yyyy-MM-dd")
        idProveedor = self.main.cbProveedor_2.currentIndex()+1
        try:
            if idProveedor:
                query.insertEncabezadoCompra(fecha, 0, 100, 1, idProveedor)
            print('error')
        except Exception as e:
            print('error '+e)
    
    def registerProduct(self):
        query = Query()
        try: 
            idProveedor = self.main.cbProveedor_2.currentIndex()+1
            for row in range(self.main.tShop.rowCount()):   
                cod = self.main.tShop.item(row, 1)
                name = self.main.tShop.item(row, 2)
                idMarca = self.main.tShop.item(row, 3)
                unidad_m = self.main.tShop.item(row, 4)
                precio_v = self.main.tShop.item(row, 6)
                query.insertProducto(cod.text(), name.text(), 0, float(precio_v.text()), unidad_m.text(), idMarca.text(), idProveedor)
        except Exception as e:
            print(e)
        
    
    def loadExcel(self,path):
        workbook = load_workbook(filename=path)
        sheet = workbook.active
        for row in sheet.iter_rows(values_only= True):
            row_index = self.main.tShop.rowCount()
            self.main.tShop.insertRow(row_index)
            print(row[1])
            self.main.tShop.setItem(row_index, 0, QTableWidgetItem(str(row[0])))#cantidad
            self.main.tShop.setItem(row_index, 1, QTableWidgetItem(str(row[1]).upper()))#codigo
            self.main.tShop.setItem(row_index, 2, QTableWidgetItem(str(row[2]).upper()))#nombre o descripcion
            self.main.tShop.setItem(row_index, 3, QTableWidgetItem(str(row[3]).upper()))#marca
            self.main.tShop.setItem(row_index, 4, QTableWidgetItem(str(row[4]).upper()))#unidad de medida
            self.main.tShop.setItem(row_index, 5, QTableWidgetItem(str(row[5])))#precio de compra
            self.main.tShop.setItem(row_index, 6, QTableWidgetItem(str(row[6])))#precio de venta          
            self.main.tShop.setItem(row_index, 7, QTableWidgetItem(str(row[7])))#precio de descuento
            self.main.tShop.setItem(row_index, 8, QTableWidgetItem(str(row[8])))#subtotal
            
    def openExcel(self):
        folder = QFileDialog()
        try: 
            folder_path, __= folder.getOpenFileName(None, 'ABRIR ARCHIVO', '', 'xlsx (*.xlsx)')
            print(folder_path)
            self.loadExcel(folder_path)
        except Exception as e:
            print(e)
          
    def registerProveedor(self):
        query = Query()
        self.nit = self.main.txtNIT.text()
        self.nombre = self.main.txtProveedor.text()
        if self.nit:
            message_box = QMessageBox(self)
            message_box.setWindowTitle("Confirmación")
            message_box.setText("¿Estás seguro de que deseas continuar?")
            message_box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            message_box.setIcon(QMessageBox.Icon.Question)
            response = message_box.exec()
            if response == QMessageBox.StandardButton.Yes:
                query.insertProveedor(self.nit, self.nombre)
                row = self.main.tProveedor.rowCount()
                self.main.tProveedor.insertRow(row)
                self.main.tProveedor.setItem(row, 0, QTableWidgetItem(self.nit))
                self.main.tProveedor.setItem(row, 1, QTableWidgetItem(self.nombre))
                self.main.txtNIT.setText("")
                self.main.txtProveedor.setText("")
                self.main.cbProveedor.clear()
                self.showProveedor()
    
    def registerMarca(self):
        query = Query()
        idProveedor = self.main.cbProveedor.currentIndex()+1
        self.marca = self.main.txtMarca.text()
        if idProveedor:
            message_box = QMessageBox(self)
            message_box.setWindowTitle("Confirmación")
            message_box.setText("¿Estás seguro de que deseas continuar?")
            message_box.setStandardButtons(QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            message_box.setIcon(QMessageBox.Icon.Question)
            response = message_box.exec()
            if response == QMessageBox.StandardButton.Yes:
                query.insertMarca(self.marca, idProveedor)
                row = self.main.tMarca.rowCount()
                self.main.tMarca.insertRow(row)
                self.main.tMarca.setItem(row, 0, QTableWidgetItem(self.marca))
                self.main.tMarca.setItem(row, 1, QTableWidgetItem(self.main.cbProveedor.currentText()))
                self.main.txtMarca.setText("")
                self.main.cbProveedor.setCurrentIndex(-1)
            
        
    def cleanTables(self):
        self.main.tProveedor.setRowCount(0)
        self.main.tMarca.setRowCount(0)